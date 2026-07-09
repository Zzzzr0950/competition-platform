"""Admin blueprint: review, statistics, catalog & user management."""

import csv
import io
from flask import (Blueprint, render_template, request, redirect,
                   url_for, session, flash, jsonify, Response, send_file)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from models import query_db, execute_db
from auth_utils import login_required, admin_required, get_current_user

admin_bp = Blueprint('admin', __name__)


@admin_bp.route('/admin')
@admin_required
def dashboard():
    """Admin dashboard with summary stats."""
    # Counts
    total = query_db("SELECT COUNT(*) as c FROM submissions", one=True)['c']
    pending = query_db(
        "SELECT COUNT(*) as c FROM submissions WHERE status = 'pending'", one=True
    )['c']
    approved_this_month = query_db("""
        SELECT COUNT(*) as c FROM submissions
        WHERE status = 'approved'
        AND reviewed_at >= datetime('now','start of month','localtime')
    """, one=True)['c']
    today = query_db("""
        SELECT COUNT(*) as c FROM submissions
        WHERE created_at >= datetime('now','start of day','localtime')
    """, one=True)['c']

    return render_template('admin/dashboard.html',
                          total=total, pending=pending,
                          approved_this_month=approved_this_month, today=today)


@admin_bp.route('/admin/review')
@admin_required
def review_list():
    """List submissions for review with filters and pagination."""
    status_filter = request.args.get('status', '').strip()
    catalog_filter = request.args.get('catalog_id', '').strip()
    class_filter = request.args.get('class_name', '').strip()
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)

    query = """
        SELECT s.*, c.name as competition_name, c.level as competition_level,
               u.name as student_name, u.student_id, u.class_name
        FROM submissions s
        JOIN competition_catalog c ON s.catalog_id = c.id
        JOIN users u ON s.user_id = u.id
        WHERE 1=1
    """
    count_query = """
        SELECT COUNT(*) as c
        FROM submissions s
        JOIN competition_catalog c ON s.catalog_id = c.id
        JOIN users u ON s.user_id = u.id
        WHERE 1=1
    """
    params = []

    if status_filter in ('pending', 'approved', 'rejected'):
        query += " AND s.status = ?"
        count_query += " AND s.status = ?"
        params.append(status_filter)
    if catalog_filter:
        query += " AND s.catalog_id = ?"
        count_query += " AND s.catalog_id = ?"
        params.append(int(catalog_filter))
    if class_filter:
        query += " AND u.class_name = ?"
        count_query += " AND u.class_name = ?"
        params.append(class_filter)
    if search:
        query += " AND (u.name LIKE ? OR u.student_id LIKE ? OR c.name LIKE ?)"
        count_query += " AND (u.name LIKE ? OR u.student_id LIKE ? OR c.name LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

    # Pagination
    per_page = 15
    total_count = query_db(count_query, params, one=True)['c']
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    offset = (page - 1) * per_page

    query += " ORDER BY s.status = 'pending' DESC, s.created_at DESC LIMIT ? OFFSET ?"
    submissions = query_db(query, params + [per_page, offset])

    # Catalog for filter dropdown
    catalog = query_db("SELECT id, name, level FROM competition_catalog WHERE is_active = 1 ORDER BY sort_order")

    return render_template('admin/review_list.html',
                          submissions=submissions, catalog=catalog,
                          page=page, total_pages=total_pages,
                          total_count=total_count,
                          status_filter=status_filter,
                          catalog_filter=catalog_filter,
                          class_filter=class_filter,
                          search=search)


@admin_bp.route('/admin/review/<int:sid>')
@admin_required
def review_detail(sid):
    """View a single submission for detailed review."""
    sub = query_db("""
        SELECT s.*, c.name as competition_name, c.level as competition_level,
               c.category as competition_category, c.is_key_competition,
               u.name as student_name, u.student_id, u.class_name,
               u.phone, u.qq
        FROM submissions s
        JOIN competition_catalog c ON s.catalog_id = c.id
        JOIN users u ON s.user_id = u.id
        WHERE s.id = ?
    """, [sid], one=True)

    if not sub:
        flash('申报记录不存在', 'error')
        return redirect(url_for('admin.review_list'))

    return render_template('admin/review_detail.html', sub=sub)


@admin_bp.route('/api/admin/review/<int:sid>/approve', methods=['POST'])
@admin_required
def approve_submission(sid):
    """Approve a submission."""
    admin = get_current_user()
    data = request.get_json() or {}
    comment = data.get('comment', '').strip()

    sub = query_db("""
        SELECT s.*, c.level as competition_level
        FROM submissions s
        JOIN competition_catalog c ON s.catalog_id = c.id
        WHERE s.id = ?
    """, [sid], one=True)
    if not sub:
        return jsonify({'error': '申报记录不存在'}), 404
    if sub['status'] != 'pending':
        return jsonify({'error': '该申报已被其他人审核过了，请刷新列表'}), 409

    old_status = sub['status']

    # Recalculate credits on approval
    from credit_calculator import calculate_credits
    credits = calculate_credits(
        sub['competition_level'],
        sub['award_tier'] or '国家级',
        sub['award_level'],
        sub['is_leader']
    )

    execute_db(
        """UPDATE submissions SET status = 'approved', review_comment = ?, credits = ?,
           reviewed_by = ?, reviewed_at = datetime('now','localtime') WHERE id = ?""",
        [comment, credits, admin['id'], sid]
    )

    action_name = 're-approved' if old_status == 'approved' else ('re-approved' if old_status == 'rejected' else 'approved')
    execute_db(
        """INSERT INTO submission_logs (submission_id, action, old_status, new_status, performed_by, comment)
           VALUES (?, ?, ?, 'approved', ?, ?)""",
        [sid, action_name, old_status, admin['id'], comment]
    )

    return jsonify({'success': True, 'message': '审核通过'})


@admin_bp.route('/api/admin/review/<int:sid>/reject', methods=['POST'])
@admin_required
def reject_submission(sid):
    """Reject a submission with required reason."""
    admin = get_current_user()
    data = request.get_json() or {}
    comment = data.get('comment', '').strip()

    if not comment:
        return jsonify({'error': '驳回时必须填写原因'}), 400

    sub = query_db("SELECT * FROM submissions WHERE id = ?", [sid], one=True)
    if not sub:
        return jsonify({'error': '申报记录不存在'}), 404
    if sub['status'] != 'pending':
        return jsonify({'error': '该申报已被其他人审核过了，请刷新列表'}), 409

    old_status = sub['status']

    execute_db(
        """UPDATE submissions SET status = 'rejected', review_comment = ?,
           reviewed_by = ?, reviewed_at = datetime('now','localtime') WHERE id = ?""",
        [comment, admin['id'], sid]
    )

    action_name = 'rejected' if old_status == 'pending' else 're-reviewed'
    execute_db(
        """INSERT INTO submission_logs (submission_id, action, old_status, new_status, performed_by, comment)
           VALUES (?, ?, ?, 'rejected', ?, ?)""",
        [sid, action_name, old_status, admin['id'], comment]
    )

    return jsonify({'success': True, 'message': '已驳回'})


@admin_bp.route('/api/admin/review/batch', methods=['POST'])
@admin_required
def batch_review():
    """Batch approve or reject submissions."""
    admin = get_current_user()
    data = request.get_json() or {}
    ids = data.get('ids', [])
    action = data.get('action', '')
    comment = data.get('comment', '').strip()

    if not ids or action not in ('approve', 'reject'):
        return jsonify({'error': '参数错误'}), 400
    if action == 'reject' and not comment:
        return jsonify({'error': '批量驳回时必须填写原因'}), 400

    new_status = 'approved' if action == 'approve' else 'rejected'
    count = 0

    for sid in ids:
        sub = query_db("SELECT * FROM submissions WHERE id = ? AND status = 'pending'", [sid], one=True)
        if sub:
            execute_db(
                """UPDATE submissions SET status = ?, review_comment = ?,
                   reviewed_by = ?, reviewed_at = datetime('now','localtime') WHERE id = ?""",
                [new_status, comment, admin['id'], sid]
            )
            execute_db(
                """INSERT INTO submission_logs (submission_id, action, old_status, new_status, performed_by, comment)
                   VALUES (?, ?, 'pending', ?, ?, ?)""",
                [sid, action + 'd', new_status, admin['id'], comment]
            )
            count += 1

    return jsonify({'success': True, 'message': f'已{new_status} {count} 条申报'})


@admin_bp.route('/admin/statistics')
@admin_required
def statistics():
    """Statistics dashboard page."""
    return render_template('admin/statistics.html')


@admin_bp.route('/api/admin/statistics/overview')
@admin_required
def statistics_overview():
    """Aggregate counts by status."""
    stats = {
        'total': query_db("SELECT COUNT(*) as c FROM submissions", one=True)['c'],
        'pending': query_db("SELECT COUNT(*) as c FROM submissions WHERE status = 'pending'", one=True)['c'],
        'approved': query_db("SELECT COUNT(*) as c FROM submissions WHERE status = 'approved'", one=True)['c'],
        'rejected': query_db("SELECT COUNT(*) as c FROM submissions WHERE status = 'rejected'", one=True)['c'],
    }
    stats['student_count'] = query_db(
        "SELECT COUNT(DISTINCT user_id) as c FROM submissions", one=True
    )['c']

    return jsonify(stats)


@admin_bp.route('/api/admin/statistics/by-level')
@admin_required
def statistics_by_level():
    """Submissions grouped by competition level."""
    rows = query_db("""
        SELECT c.level, COUNT(*) as count
        FROM submissions s
        JOIN competition_catalog c ON s.catalog_id = c.id
        GROUP BY c.level
        ORDER BY c.level
    """)
    return jsonify([{'level': r['level'], 'count': r['count']} for r in rows])


@admin_bp.route('/api/admin/statistics/by-competition')
@admin_required
def statistics_by_competition():
    """Top competitions by submission count."""
    limit = request.args.get('limit', 10, type=int)
    rows = query_db("""
        SELECT c.name, c.level, COUNT(*) as count
        FROM submissions s
        JOIN competition_catalog c ON s.catalog_id = c.id
        GROUP BY s.catalog_id
        ORDER BY count DESC
        LIMIT ?
    """, [limit])
    return jsonify([{
        'name': r['name'],
        'level': r['level'],
        'count': r['count']
    } for r in rows])


@admin_bp.route('/api/admin/classes')
@admin_required
def all_classes():
    """All student classes (including those with 0 submissions)."""
    rows = query_db("""
        SELECT class_name, COUNT(*) as total
        FROM users WHERE role = 'student'
        GROUP BY class_name ORDER BY class_name
    """)
    return jsonify([{'class_name': r['class_name'], 'total': r['total']} for r in rows])


@admin_bp.route('/api/admin/statistics/by-class')
@admin_required
def statistics_by_class():
    """Submissions grouped by student class."""
    rows = query_db("""
        SELECT u.class_name, COUNT(*) as count
        FROM submissions s
        JOIN users u ON s.user_id = u.id
        GROUP BY u.class_name
        ORDER BY count DESC
    """)
    return jsonify([{'class_name': r['class_name'], 'count': r['count']} for r in rows])


@admin_bp.route('/api/admin/export')
@admin_required
def export_csv():
    """Export submissions as CSV file, optionally filtered by class.
    Same student's 学号/姓名/班级/总学分 merged into one cell (written once).
    """
    class_name = request.args.get('class_name', '').strip()

    # LEFT JOIN so students with 0 submissions still appear
    sql = """
        SELECT u.id as uid, u.student_id, u.name as student_name, u.class_name,
               c.name as competition_name, c.level as competition_level,
               s.award_tier, s.award_level, s.award_date, s.team_name, s.team_members,
               s.status, s.review_comment, s.created_at, s.reviewed_at,
               s.credits,
               COALESCE((SELECT SUM(s2.credits) FROM submissions s2
                          WHERE s2.user_id = u.id AND s2.status = 'approved'), 0) as total_credits
        FROM users u
        LEFT JOIN submissions s ON s.user_id = u.id
        LEFT JOIN competition_catalog c ON s.catalog_id = c.id
        WHERE u.role = 'student'
    """
    params = []

    if class_name:
        sql += " AND u.class_name = ?"
        params.append(class_name)

    sql += " ORDER BY u.student_id ASC"

    rows = query_db(sql, params)

    # Build Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = '竞赛获奖导出'

    # Styles
    header_fill = PatternFill(start_color='1A56DB', end_color='1A56DB', fill_type='solid')
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    cell_font = Font(name='微软雅黑', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    # Header
    headers = ['学号', '姓名', '竞赛名称', '竞赛级别', '获奖层次',
               '获奖等级', '获奖日期', '团队名称', '团队成员',
               '审核状态', '审核意见', '提交时间', '审核时间',
               '学分', '总学分']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Column widths
    widths = [16, 10, 36, 10, 10, 18, 14, 18, 22, 10, 22, 20, 20, 8, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Group by student
    student_groups = {}
    student_order = []
    for r in rows:
        uid = r['uid']
        if uid not in student_groups:
            student_groups[uid] = {
                'student_id': r['student_id'],
                'name': r['student_name'],
                'total_credits': r['total_credits'],
                'submissions': []
            }
            student_order.append(uid)
        student_groups[uid]['submissions'].append(r)

    # Write data rows
    status_map = {'pending': '待审核', 'approved': '已通过', 'rejected': '已驳回'}
    row_num = 2

    for uid in student_order:
        stu = student_groups[uid]
        subs = stu['submissions']
        start_row = row_num
        end_row = row_num + len(subs) - 1

        for idx, r in enumerate(subs):
            created = r['created_at'] or ''
            reviewed = r['reviewed_at'] or ''

            vals = [
                r['student_id'],
                stu['name'],
                r['competition_name'],
                r['competition_level'],
                r['award_tier'] or '',
                r['award_level'],
                r['award_date'],
                r['team_name'],
                r['team_members'],
                status_map.get(r['status'], r['status']),
                r['review_comment'],
                created.replace(' ', 'T') if created else '',
                reviewed.replace(' ', 'T') if reviewed else '',
                r['credits'],
                stu['total_credits'],
            ]

            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = left_align if col_idx in (3, 11) else center_align

            row_num += 1

        # Merge cells for 学号/姓名/总学分
        if len(subs) > 1:
            for merge_col in [1, 2, 15]:
                ws.merge_cells(start_row=start_row, start_column=merge_col,
                              end_row=end_row, end_column=merge_col)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'competition_export{("_" + class_name) if class_name else ""}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ===== Competition Catalog Management =====

@admin_bp.route('/admin/catalog')
@admin_required
def manage_catalog():
    """Manage competition catalog entries."""
    catalog = query_db("""
        SELECT * FROM competition_catalog ORDER BY sort_order
    """)
    return render_template('admin/manage_catalog.html', catalog=catalog)


@admin_bp.route('/api/admin/catalog', methods=['POST'])
@admin_required
def add_catalog():
    """Add a new catalog entry."""
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    short_name = data.get('short_name', '').strip()
    category = data.get('category', '学科竞赛').strip()
    level = data.get('level', '校级').strip()
    organizer = data.get('organizer', '').strip()

    if not name:
        return jsonify({'error': '竞赛名称不能为空'}), 400

    cid = execute_db(
        """INSERT INTO competition_catalog (name, short_name, category, level, organizer)
           VALUES (?, ?, ?, ?, ?)""",
        [name, short_name, category, level, organizer]
    )
    return jsonify({'success': True, 'id': cid, 'message': '已添加'})


@admin_bp.route('/api/admin/catalog/<int:cid>', methods=['PUT'])
@admin_required
def update_catalog(cid):
    """Update a catalog entry."""
    data = request.get_json() or {}
    execute_db(
        """UPDATE competition_catalog SET name=?, short_name=?, category=?, level=?,
           organizer=?, sort_order=? WHERE id=?""",
        [data.get('name', ''), data.get('short_name', ''),
         data.get('category', ''), data.get('level', ''),
         data.get('organizer', ''),
         data.get('sort_order', 0), cid]
    )
    return jsonify({'success': True, 'message': '已更新'})


@admin_bp.route('/api/admin/catalog/<int:cid>', methods=['DELETE'])
@admin_required
def delete_catalog(cid):
    """Soft-delete a catalog entry."""
    # Check if any submissions reference this
    count = query_db(
        "SELECT COUNT(*) as c FROM submissions WHERE catalog_id = ?", [cid], one=True
    )['c']
    if count > 0:
        # Soft delete
        execute_db("UPDATE competition_catalog SET is_active = 0 WHERE id = ?", [cid])
        return jsonify({'success': True, 'message': f'已禁用（关联 {count} 条申报记录，已保留数据）'})
    else:
        execute_db("DELETE FROM competition_catalog WHERE id = ?", [cid])
        return jsonify({'success': True, 'message': '已删除'})


# ===== User Management =====

@admin_bp.route('/admin/users')
@admin_required
def user_management():
    """List all student users with submission counts (single query)."""
    search = request.args.get('search', '').strip()
    query = """
        SELECT u.*, COALESCE(s.submission_count, 0) as submission_count
        FROM users u
        LEFT JOIN (
            SELECT user_id, COUNT(*) as submission_count
            FROM submissions GROUP BY user_id
        ) s ON u.id = s.user_id
        WHERE u.role = 'student'
    """
    params = []

    if search:
        query += " AND (u.name LIKE ? OR u.student_id LIKE ? OR u.class_name LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

    query += " ORDER BY u.created_at DESC"
    users = query_db(query, params)

    return render_template('admin/user_management.html', users=users, search=search)


@admin_bp.route('/api/admin/users/<int:uid>/toggle-active', methods=['POST'])
@admin_required
def toggle_user_active(uid):
    """Enable or disable a user account."""
    user = query_db("SELECT * FROM users WHERE id = ? AND role = 'student'", [uid], one=True)
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    new_status = 0 if user['is_active'] else 1
    execute_db("UPDATE users SET is_active = ? WHERE id = ?", [new_status, uid])

    action = '启用' if new_status else '禁用'
    return jsonify({'success': True, 'message': f'已{action}用户 {user["name"]}'})


@admin_bp.route('/api/admin/users/<int:uid>/reset-password', methods=['POST'])
@admin_required
def reset_user_password(uid):
    """Reset a student's password to 123456 and force change on next login."""
    from auth_utils import hash_password

    user = query_db("SELECT * FROM users WHERE id = ? AND role = 'student'", [uid], one=True)
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    execute_db(
        "UPDATE users SET password_hash = ?, must_change_password = 1 WHERE id = ?",
        [hash_password('123456'), uid]
    )
    return jsonify({'success': True, 'message': f'已重置 {user["name"]} 的密码为 123456，下次登录需修改'})
