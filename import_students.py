"""批量导入学生账号 — 支持 Excel(.xlsx) 和 CSV
使用方法：
  python import_students.py 学生名单.xlsx
  python import_students.py 学生名单.xlsx 123456    （指定统一初始密码）
  python import_students.py 学生名单.csv

Excel/CSV 需要包含列：学号、姓名、班级、专业
（列名支持模糊匹配：学号/学生号、姓名/名字、班级/班、专业/专业名称）
"""
import os
import sys
from auth_utils import hash_password

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database', 'platform.db')


def find_column(headers, keywords):
    """从表头中模糊匹配列名。"""
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw in str(h):
                return i
    return None


def import_students(file_path, default_password=None):
    """导入学生账号。自动识别 Excel 或 CSV。"""
    import sqlite3

    if not os.path.exists(file_path):
        print(f'[错误] 文件不存在: {file_path}')
        return

    ext = os.path.splitext(file_path)[1].lower()

    # ── 读取数据 ──
    rows = []
    if ext in ('.xlsx', '.xls'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            print('[错误] 需要安装 openpyxl: pip install openpyxl')
            return
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        headers = [str(c.value or '') for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append([str(v or '').strip() for v in row])
        wb.close()
    elif ext == '.csv':
        import csv
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            headers = [str(h).strip() for h in next(reader)]
            rows = [[str(v or '').strip() for v in row] for row in reader]
    else:
        print(f'[错误] 不支持的文件格式: {ext}，请用 .xlsx 或 .csv')
        return

    # ── 定位列 ──
    idx_id = find_column(headers, ['学号', '学生号', 'student_id', '学籍号'])
    idx_name = find_column(headers, ['姓名', '名字', 'name', '学生姓名'])
    idx_class = find_column(headers, ['班级', '班', 'class', '班级名称'])
    idx_major = find_column(headers, ['专业', '专业名称', 'major', '专业方向'])

    if idx_id is None or idx_name is None:
        print(f'[错误] 未找到"学号"或"姓名"列，表头: {headers}')
        return
    if idx_class is None:
        print(f'[警告] 未找到"班级"列，将留空')
    if idx_major is None:
        print(f'[警告] 未找到"专业"列，将留空')

    print(f'检测到列: 学号=第{idx_id+1}列, 姓名=第{idx_name+1}列, '
          f'班级={"第"+str(idx_class+1)+"列" if idx_class is not None else "未找到"}, '
          f'专业={"第"+str(idx_major+1)+"列" if idx_major is not None else "未找到"}')
    print(f'共 {len(rows)} 行数据，开始导入...\n')

    # ── 写入数据库 ──
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    created = 0
    skipped = 0
    class_stats = {}  # 班级 -> 人数

    for i, row in enumerate(rows, start=2):
        student_id = row[idx_id] if idx_id < len(row) else ''
        name = row[idx_name] if idx_name < len(row) else ''
        class_name = row[idx_class] if idx_class is not None and idx_class < len(row) else ''
        major = row[idx_major] if idx_major is not None and idx_major < len(row) else ''

        if not student_id or not name:
            skipped += 1
            continue

        # 检查已存在
        existing = cursor.execute(
            "SELECT id FROM users WHERE student_id = ?", [student_id]
        ).fetchone()
        if existing:
            skipped += 1
            continue

        pwd = default_password if default_password else (student_id[-6:] if len(student_id) >= 6 else student_id)

        try:
            cursor.execute(
                """INSERT INTO users (student_id, name, class_name, major, password_hash, must_change_password)
                   VALUES (?, ?, ?, ?, ?, 1)""",
                [student_id, name, class_name, major, hash_password(pwd)]
            )
            created += 1
            class_stats[class_name] = class_stats.get(class_name, 0) + 1
        except Exception as e:
            print(f'  [跳过] 第{i}行 {student_id} {name}: {e}')
            skipped += 1

    conn.commit()
    conn.close()

    # ── 结果 ──
    print(f'===== 导入完成 =====')
    print(f'新增: {created} 人')
    print(f'跳过（已存在/数据不全）: {skipped} 人')
    print(f'初始密码: {default_password if default_password else "学号后6位"}')
    print(f'首次登录强制修改密码')
    if class_stats:
        print(f'\n各班导入人数:')
        for cls in sorted(class_stats.keys()):
            print(f'  {cls}: {class_stats[cls]} 人')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python import_students.py <文件.xlsx|.csv> [默认密码]')
        print('示例: python import_students.py students.xlsx')
        print('      python import_students.py students.xlsx 123456')
        sys.exit(1)

    pwd = sys.argv[2] if len(sys.argv) > 2 else None
    import_students(sys.argv[1], pwd)
